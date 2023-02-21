from datetime import datetime
from openpyxl import load_workbook

Rut="C:\\Users\\SENA\\OneDrive - Universidad Nacional Abierta y a Distancia\\Desktop\\wass\\Repositorio_prueba\\Crud_R_epo.xlsx"
Rut=r"C:\Users\SENA\OneDrive - Universidad Nacional Abierta y a Distancia\Desktop\wass\Repositorio_prueba\Crud_R_epo.xlsx"

def leer(ruta:str, extraer:str):
    Archivo_Exccel = load_workbook(ruta)
    Hoja_datos= Archivo_Exccel['Hoja_repa']
    Hoja_datos=Hoja_datos['A2':'E'+str(Hoja_datos.max_row)]

    info={}

    for i in Hoja_datos:
        
        if isinstance(i[0].value, int):
            info.setdefault(i[0].value,{'Nombre':i[1].value, 'Categoria':i[2].value, 'Precio':i[3].value, 'Cantidad':i[4].value})

    if not(extraer=='todo'):
        info=filtrar(info, extraer)

    for i in info:
        print('***  Nombre  **')  
        print('id: '+str(i)+'\n'+'Nombre: '+str(info[i]['Nombre'])+'\n'+'Categoria: '+str(info[i]['Categoria'])+'\n'+'Precio'+str(info[i]['Precio']) +'\n'+'Cantidad'+str(info[i]['Cantidad'])             )  
        print()
    return          

def filtrar(info:dict, filtro:str):
    aux={}
    for i in info:
        if info[i]['Categoria']==filtro:
            aux.setdefault(i, info[i])
    return aux

def actualizar(ruta:str, identificador:int, datos_actualizados:dict):
    Archivo_Exccel=load_workbook(ruta)
    Hoja_datos =Archivo_Exccel['Hoja_repa']
    Hoja_datos=Hoja_datos['A2':'E'+str(Hoja_datos.max_row)]
    hoja=Archivo_Exccel.active

    Nombre=2
    Categoria=3
    Precio=4
    cantidad=5
    encontro=False
    for i in Hoja_datos:
        if i[0].value==identificador:
            fila=i[0].row
            encontro=True
            for d in datos_actualizados:
                if d=='Nombre' and not (datosActualizados[d]==''):
                    hoja.cell(row=fila, column=Nombre).value=datosActualizados[d]
                elif d=='Categoria' and not (datosActualizados[d]==''):
                    hoja.cell(row=fila, column=Categoria).value=datosActualizados[d] 
                elif d=='Precio' and not (datosActualizados[d]==''):
                    hoja.cell(row=fila, column=Precio).value=datosActualizados[d] 
                elif d=='Cantidad' and not (datosActualizados[d]==''):
                    hoja.cell(row=fila, column=cantidad).value=datosActualizados[d] 
    Archivo_Exccel.save(ruta)  
    if encontro==False:
        print('Error: No existe una tarea con ese ID')
        print()
    return

def agregar(ruta:int, datos:dict):
    Archivo_Exccel=load_workbook(ruta)
    Hoja_datos=Archivo_Exccel['Hoja_repa']
    Hoja_datos=Hoja_datos['A2':'F'+str(Hoja_datos.max_row+1)]                                 
    hoja=Archivo_Exccel.active

    Nombre=2
    Categoria=3
    Precio=4
    Cantidad=5
    for i in Hoja_datos:
        if not(isinstance(i[0].value, int )):
            identificador=i[0].row
            hoja.cell(row=identificador, column=1).value=identificador-1
            hoja.cell(row=identificador, column=Nombre).value=datos['Nombre']
            hoja.cell(row=identificador, column=Categoria).value=datos['Categoria']
            hoja.cell(row=identificador, column=Precio).value=datos['Precio']
            hoja.cell(row=identificador, column=Cantidad).value=datos['Cantidad']
            break
    Archivo_Exccel.save(ruta)
    return


def borrar(ruta, identificador):
    Archivo_Exccel= load_workbook(ruta)
    Hoja_datos=Archivo_Exccel['Hoja_repa']
    Hoja_datos=Hoja_datos['A2':'E'+str(Hoja_datos.max_row)]
    hoja=Archivo_Exccel.active

    Nombre=2
    Categoria=3
    Precio=4
    Cantidad=5
    encontro=False
    for i in Hoja_datos:
        if i[0].value==identificador:
            fila=i[0].row
            encontro=True

            hoja.cell(row=fila, column=1).value=""
            hoja.cell(row=fila, column=Nombre).value=""
            hoja.cell(row=fila, column=Categoria).value=""
            hoja.cell(row=fila, column=Precio).value=""
            hoja.cell(row=fila, column=Cantidad).value=""
    Archivo_Exccel.save(ruta) 
    if encontro==False:
        print('Error: no exixte una tarea con ese Id')
        print()
    return           
           
Rut="C:\\Users\\SENA\\OneDrive - Universidad Nacional Abierta y a Distancia\\Desktop\\wass\\Repositorio_prueba\\Crud_R_epo.xlsx"

datosActualizados={'Nombre':'', 'Categoria':'', 'Precio':'', 'Cantidad':''}
while True:
    print('indique la accion que desea realizar: ')
    print('Consultar: 1')
    print('Actualizar: 2')
    print('Crear nueva Categoria: 3')
    print('Borrar: 4')
    accion= input('Escriba la opcion: ')
    if not(accion=='1') and not (accion=='2') and not (accion=='3') and not (accion=='4'):
        print('Comando invalido por favor eliga una opcion valida')
    elif accion=='1':
        opc_consulta=''
        print('Indique la Categoria que desea consultar: ')
        print('Todas las Categorias: 1')
        print('Evaluado: 2')
        print('En Venta: 3')
        print('Por Comprar: 4')
        print('Agotado: 5')
        opc_consulta= input('Escriba la Categoria que desea consultar:')
        if opc_consulta=='1':
            print()
            print()
            print('*** Consultando todas las Categorias ***')
            leer(Rut,'todo')
        elif opc_consulta=='2':
            print()
            print()
            print('*** Consultando Categoria Evaluada ***')
            leer(Rut,'Evaluada')
        elif opc_consulta=='3':
            print()
            print()
            print('*** Consultando Categoria En Venta ***')
            leer(Rut,'En Venta')
        elif opc_consulta=='4':
            print()
            print()
            print('*** Consultando Categoria Por Comprar ***')
            leer(Rut,'Por Comprar')
        elif opc_consulta=='5':
            print() 
            print()
            print('*** Consultando Categoria Agotada ***')
            leer(Rut,'Agotada')

    elif accion=='2':
        datosActualizados={'Nombre':'', 'Categoria':'', 'Precio':'', 'Cantidad':''}
        print('*** Actualizar Producto ***')
        print()
        id_Actualizar=int(input('Indique el ID del Producto que desea actualizar: '))
        print()
        print('** Nuevo Nombre **')
        print('*** Nota: Si no desea actualizar el Nombre solo oprima ENTER ***')
        datosActualizados['Nombre']=input('Indique el nuevo Nombre de la tarea: ')
        print()
        print('** Nueva Categoria **')
        datosActualizados['Categoria']=input('Indique la nueva Categoria de la tarea: ')
        print()
        print('** Nueva Categoria **')
        print('Evaluado: 2')
        print('En Venta: 3')
        print('Por Comprar: 4')
        print('Agotado: 5')
        print('*** Nota: si no desea actualizar la Categoria oprima ENTER')
        estadoNuevo=input('Indique el nuevo estado de la tarea: ')
        if estadoNuevo=='2':
            datosActualizados['Categoria']='Evaluado'
        if estadoNuevo=='3':
            datosActualizados['Categoria']='En Venta'
        if estadoNuevo=='4':
            datosActualizados['Categoria']='Por Comprar'
        if estadoNuevo=='5':
            now=datetime.now()
            datosActualizados['Categoria']='Agotada'

        now=datetime.now()
        datosActualizados['fecha inico']=str(now.day)+'/'+str(now.month)+'/'+str(now.year)
        actualizar(Rut,id_Actualizar, datosActualizados)
        print()

    elif accion=='3':
        datosActualizados={'Nombre':'', 'Categoria':'', 'Precio':'', 'Cantidad':''}
        print('** Crear nuevo Nombre **')
        print()
        print('** Nombre **')
        print()
        datosActualizados['Nombre']=input('Indique el Nombre del Producto: ')
        print()
        print('** Nueva Categoria **')
        datosActualizados['Categoria']=input('Indique la nueva Categoria de la tarea: ')
        print()
        print('** Nueva Categoria **')
        print('Evaluado: 2')
        print('En Venta: 3')
        print('Por Comprar: 4')
        print('Agotado: 5')
        print('*** Nota: si no desea actualizar la Categoria oprima ENTER')
        estadoNuevo=input('Indique la nueva Categoria del Producto: ')
        if estadoNuevo=='2':
            datosActualizados['Categoria']='Evaluado'
        if estadoNuevo=='3':
            datosActualizados['Categoria']='En Venta'
        if estadoNuevo=='4':
            datosActualizados['Categoria']='Por Comprar'
        if estadoNuevo=='5':
            now=datetime.now()
            datosActualizados['Categoria']='Agotada'
        print('** Indique El Precio **')
        print()
        print('** Precio **')
        datosActualizados['Precio']=input('indique el precio del producto : ')
        print('** Indique La Cantidad **')
        print()
        print('** Cantidad **')
        datosActualizados['Cantidad']= input('Indique La Cantidad de producto: ')
        now=datetime.now()
        datosActualizados['fecha inicio']=str(now.day)+'/'+str(now.month)+'/'+str(now.year)
        datosActualizados['fecha finalizacion']=''
        agregar(Rut, datosActualizados)
    elif accion=='4':
        print('')
        print('** Eliminar Nombre**')
        iden=int(input('Indique el ID del Nombre que desea eliminar : '))  
        borrar(Rut,iden)  